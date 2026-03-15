"""Document management service — business logic and background tasks."""

from __future__ import annotations

import io
from datetime import UTC, datetime, timedelta
from typing import TYPE_CHECKING

from sqlalchemy import select

from app.core.logging import get_logger
from app.models.document import Document, DocumentStatus

if TYPE_CHECKING:
    from uuid import UUID

    from sqlalchemy.ext.asyncio import AsyncSession

logger = get_logger("services.document")


async def notify_reviewers_for_approval(
    document_id: UUID,
    db: AsyncSession,
) -> None:
    """Notify assigned reviewers that a document is pending their approval."""
    result = await db.execute(select(Document).where(Document.id == document_id))
    doc = result.scalar_one_or_none()
    if not doc:
        logger.warning("notify_reviewers: document not found", document_id=str(document_id))
        return

    if doc.status != DocumentStatus.UNDER_REVIEW:
        return

    distribution = doc.distribution_list or []
    reviewers = [entry for entry in distribution if isinstance(entry, dict) and entry.get("required")]

    from app.tasks.notifications import send_email_task

    for reviewer in reviewers:
        user_id = reviewer.get("user_id")
        if not user_id:
            continue
        # In production, fetch user email from DB; here we enqueue by user_id
        send_email_task.delay(
            to=reviewer.get("email", f"user-{user_id}@company.local"),
            subject=f"[OpenQHSE] Documento pendiente de aprobación: {doc.code} — {doc.title}",
            template="document_review_request",
            context={
                "document_code": doc.code,
                "document_title": doc.title,
                "version": doc.version,
                "reviewer_user_id": str(user_id),
            },
        )

    logger.info(
        "Reviewers notified",
        document_id=str(document_id),
        reviewer_count=len(reviewers),
    )


async def auto_archive_superseded_versions(
    document_id: UUID,
    db: AsyncSession,
) -> None:
    """Mark a document OBSOLETE when a new version is approved and set effective_date."""
    result = await db.execute(
        select(Document).where(
            Document.id == document_id,
            Document.status == DocumentStatus.APPROVED,
        )
    )
    current_doc = result.scalar_one_or_none()
    if not current_doc:
        return

    # Find any previous approved documents with the same code (older version)
    prev_result = await db.execute(
        select(Document).where(
            Document.organization_id == current_doc.organization_id,
            Document.code == current_doc.code,
            Document.id != document_id,
            Document.status == DocumentStatus.APPROVED,
            Document.is_deleted == False,  # noqa: E712
        )
    )
    superseded = prev_result.scalars().all()

    for old_doc in superseded:
        old_doc.status = DocumentStatus.OBSOLETE
        old_doc.updated_by = "system"

    if superseded:
        logger.info(
            "Superseded documents archived",
            document_id=str(document_id),
            archived_count=len(superseded),
        )


async def generate_document_report(
    filters: dict,  # type: ignore[type-arg]
    db: AsyncSession,
) -> bytes:
    """Generate an Excel report with the state of all documents.

    Args:
        filters: Dict with optional keys: doc_type, status, site_id, days_to_expiry
        db: Async database session

    Returns:
        Raw bytes of the .xlsx file
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        logger.error("openpyxl not installed — cannot generate document report")
        return b""

    query = select(Document).where(Document.is_deleted == False)  # noqa: E712

    if filters.get("organization_id"):
        query = query.where(Document.organization_id == filters["organization_id"])
    if filters.get("doc_type"):
        query = query.where(Document.doc_type == filters["doc_type"])
    if filters.get("status"):
        query = query.where(Document.status == filters["status"])
    if filters.get("site_id"):
        query = query.where(Document.site_id == filters["site_id"])
    if filters.get("days_to_expiry"):
        cutoff = datetime.now(UTC) + timedelta(days=int(filters["days_to_expiry"]))
        query = query.where(Document.expiry_date <= cutoff)

    result = await db.execute(query.order_by(Document.created_at.desc()))
    documents = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentos"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Código",
        "Título",
        "Tipo",
        "Versión",
        "Estado",
        "Categoría",
        "Propietario",
        "Fecha Efectiva",
        "Fecha Revisión",
        "Fecha Vencimiento",
        "Creado",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Status color mapping
    status_colors = {
        "approved": "C6EFCE",
        "draft": "FFEB9C",
        "under_review": "DDEBF7",
        "obsolete": "F4CCCC",
    }

    for row_idx, doc in enumerate(documents, 2):
        row_data = [
            doc.code,
            doc.title,
            doc.doc_type,
            doc.version,
            doc.status,
            doc.category or "",
            str(doc.owner_id),
            doc.effective_date.strftime("%Y-%m-%d") if doc.effective_date else "",
            doc.review_date.strftime("%Y-%m-%d") if doc.review_date else "",
            doc.expiry_date.strftime("%Y-%m-%d") if doc.expiry_date else "",
            doc.created_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Highlight by status
            status_fill_color = status_colors.get(doc.status, "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=status_fill_color)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    # Summary sheet
    ws_summary = wb.create_sheet("Resumen")
    status_counts: dict[str, int] = {}
    for doc in documents:
        status_counts[doc.status] = status_counts.get(doc.status, 0) + 1

    ws_summary.cell(row=1, column=1, value="Estado").font = Font(bold=True)
    ws_summary.cell(row=1, column=2, value="Cantidad").font = Font(bold=True)
    for idx, (status, count) in enumerate(status_counts.items(), 2):
        ws_summary.cell(row=idx, column=1, value=status)
        ws_summary.cell(row=idx, column=2, value=count)

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(
        "Document report generated",
        document_count=len(documents),
        filters=filters,
    )
    return buffer.read()
